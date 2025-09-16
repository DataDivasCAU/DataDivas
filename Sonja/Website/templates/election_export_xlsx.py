# export_xlsx.py
from __future__ import annotations
from io import BytesIO
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.chart import PieChart, Reference, ScatterChart, Series
from openpyxl.chart.series import DataPoint

# ---------- Farben & Helfer ----------
PARTY_COLORS = {
    "cdu": "000000", "csu": "000000",
    "spd": "E3000F",
    "fdp": "FFED00",
    "gruene": "1AA037", "grüne": "1AA037",
    "die linke": "BE3075", "linke": "BE3075",
    "afd": "009EE0",
    "bsw": "00B3A4",
    "sonstige": "9CA3AF",
}

def _norm(s: str) -> str:
    return str(s).strip().lower()

def _pretty(p: str) -> str:
    m = {
        "gruene": "Grüne", "grüne": "Grüne", "linke": "Linke",
        "cdu": "CDU", "csu": "CSU", "spd": "SPD", "fdp": "FDP",
        "afd": "AfD", "bsw": "BSW"
    }
    return m.get(_norm(p), p)

def _party_color_hex(party: str) -> str:
    k = _norm(party)
    if k in PARTY_COLORS: return PARTY_COLORS[k]
    if "gruen" in k or "grün" in k: return PARTY_COLORS["gruene"]
    if "linke" in k: return PARTY_COLORS["linke"]
    if "cdu" in k or "csu" in k: return PARTY_COLORS["cdu"]
    if "spd" in k: return PARTY_COLORS["spd"]
    if "fdp" in k: return PARTY_COLORS["fdp"]
    if "afd" in k: return PARTY_COLORS["afd"]
    if "bsw" in k: return PARTY_COLORS["bsw"]
    return PARTY_COLORS["sonstige"]

def _contrast_font(hexcolor: str) -> Font:
    c = hexcolor.lstrip("#")
    r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
    luma = 0.2126*r + 0.7152*g + 0.0722*b
    return Font(color="FFFFFF" if luma < 140 else "000000")

def _color_pie_by_party(pie: PieChart, labels: List[str]) -> None:
    # färbt die Pie-Segmente in Parteifarben (Reihenfolge = Kategorienreihenfolge)
    dps: List[DataPoint] = []
    for idx, label in enumerate(labels):
        dp = DataPoint(idx=idx)
        dp.graphicalProperties.solidFill = _party_color_hex(label)
        dps.append(dp)
    if pie.series:
        pie.series[0].data_points = dps


# ---------- Normalisierung ----------
def _rows_from_posts_section(sec: Any) -> List[dict]:
    """
    Akzeptiert Liste [{party,posts}] oder Dict {party:posts}.
    Liefert Liste von Rows [{party, posts}].
    """
    rows: List[dict] = []
    if isinstance(sec, list):
        for r in sec:
            p = _pretty(r.get("party", ""))
            rows.append({"party": p, "posts": float(r.get("posts", 0) or 0)})
    elif isinstance(sec, dict):
        for k, v in sec.items():
            rows.append({"party": _pretty(k), "posts": float(v or 0)})
    return rows

def _map_from_election_section(sec: Any) -> Dict[str, float]:
    """
    Akzeptiert Liste mit einem Dict oder reines Dict; liefert {party: %}.
    """
    if isinstance(sec, list):
        obj = (sec[0] if sec else {}) or {}
    elif isinstance(sec, dict):
        obj = sec
    else:
        obj = {}
    return { _pretty(k): float(v or 0) for k, v in obj.items() }

def _add_election(rows: List[dict], emap: Dict[str, float]) -> List[dict]:
    parties = { _pretty(r["party"]) for r in rows } | { _pretty(k) for k in emap.keys() }
    out: List[dict] = []
    for p in sorted(parties):
        base = next((x for x in rows if _pretty(x["party"]) == p), {"party": p, "posts": 0})
        out.append({
            "party": p,
            "posts": float(base.get("posts", 0)),
            "election": float(emap.get(p, 0))
        })
    return out


# ---------- Hauptfunktion ----------
def build_workbook(posts_data: dict, election_data: dict) -> BytesIO:
    """
    posts_data:  { "firstElection": {..} | [..],
                   "secondElection": {..} | [..],
                   "diffPost": {...} (optional) }
    election_data: { "election2025": {...} | [{...}],
                     "election2021": {...} | [{...}] }
    """
    # Posts normalisieren
    first_rows  = _rows_from_posts_section(posts_data.get("firstElection", {}))
    second_rows = _rows_from_posts_section(posts_data.get("secondElection", {}))

    # Election (%)
    e25_map = _map_from_election_section(election_data.get("election2025", {}))
    e21_map = _map_from_election_section(election_data.get("election2021", {}))

    # Mergen (Tabellen zeigen auch die Wahlprozente; Pies zeigen Posts)
    first  = _add_election(first_rows,  e25_map)
    second = _add_election(second_rows, e21_map)

    # Diff-Posts (optional aus JSON, sonst second-first)
    diff_map_raw = posts_data.get("diffPost") or {}
    diff_map: Dict[str, float] = { _pretty(k): float(v or 0) for k, v in (diff_map_raw or {}).items() }
    if not diff_map:
        _first_map  = { r["party"]: r["posts"] for r in first }
        _second_map = { r["party"]: r["posts"] for r in second }
        parties_all = set(_first_map) | set(_second_map)
        diff_map = { p: _second_map.get(p, 0.0) - _first_map.get(p, 0.0) for p in parties_all }

    # -------- Excel aufbauen --------
    wb = Workbook()
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True)
    header_style.alignment = Alignment(vertical="center")
    if "header_style" not in wb.named_styles:
        wb.add_named_style(header_style)

    # ===== Sheet 2025 =====
    ws25 = wb.active
    ws25.title = "2025"
    ws25.append(["Partei", "Posts", "Election (%)"])
    for c in ws25[1]: c.style = header_style
    for r in first:
        ws25.append([r["party"], r["posts"], r["election"] / 100.0])  # echte % für Excel
    ws25.column_dimensions["A"].width, ws25.column_dimensions["B"].width, ws25.column_dimensions["C"].width = 18, 12, 16
    for i in range(2, ws25.max_row + 1):
        ws25.cell(i, 3).number_format = "0.0%"
        bg = _party_color_hex(ws25.cell(i, 1).value)
        ws25.cell(i, 1).fill = PatternFill("solid", fgColor=bg)
        ws25.cell(i, 1).font = _contrast_font(bg)

    # Pie = POSTS (Spalte 2) – farblich je Partei
    pie25 = PieChart()
    pie25.title = "Posts 2025"
    pie25.add_data(Reference(ws25, min_col=2, min_row=1, max_row=ws25.max_row), titles_from_data=True)
    pie25.set_categories(Reference(ws25, min_col=1, min_row=2, max_row=ws25.max_row))
    _color_pie_by_party(pie25, [ws25.cell(r, 1).value for r in range(2, ws25.max_row + 1)])
    ws25.add_chart(pie25, "E2")

    # ===== Sheet 2021 =====
    ws21 = wb.create_sheet("2021")
    ws21.append(["Partei", "Posts", "Election (%)"])
    for c in ws21[1]: c.style = header_style
    for r in second:
        ws21.append([r["party"], r["posts"], r["election"] / 100.0])
    ws21.column_dimensions["A"].width, ws21.column_dimensions["B"].width, ws21.column_dimensions["C"].width = 18, 12, 16
    for i in range(2, ws21.max_row + 1):
        ws21.cell(i, 3).number_format = "0.0%"
        bg = _party_color_hex(ws21.cell(i, 1).value)
        ws21.cell(i, 1).fill = PatternFill("solid", fgColor=bg)
        ws21.cell(i, 1).font = _contrast_font(bg)

    # Pie = POSTS (Spalte 2) – farblich je Partei
    pie21 = PieChart()
    pie21.title = "Posts 2021"
    pie21.add_data(Reference(ws21, min_col=2, min_row=1, max_row=ws21.max_row), titles_from_data=True)
    pie21.set_categories(Reference(ws21, min_col=1, min_row=2, max_row=ws21.max_row))
    _color_pie_by_party(pie21, [ws21.cell(r, 1).value for r in range(2, ws21.max_row + 1)])
    ws21.add_chart(pie21, "E2")

    # ===== Sheet Both (mit Scatter statt Pie) =====
    wsB = wb.create_sheet("Both")
    wsB.append([
        "Partei", "First Posts", "First Election (%)",
        "Second Posts", "Second Election (%)", "Diff Posts (2-1)"
    ])
    for c in wsB[1]: c.style = header_style

    pf_posts  = {r["party"]: r["posts"]    for r in first}
    pf_elec   = {r["party"]: r["election"] for r in first}
    ps_posts  = {r["party"]: r["posts"]    for r in second}
    ps_elec   = {r["party"]: r["election"] for r in second}
    parties   = sorted(set(pf_posts) | set(ps_posts) | set(pf_elec) | set(ps_elec) | set(diff_map))

    for p in parties:
        wsB.append([
            p,
            pf_posts.get(p, 0.0), (pf_elec.get(p, 0.0)) / 100.0,
            ps_posts.get(p, 0.0), (ps_elec.get(p, 0.0)) / 100.0,
            diff_map.get(p, 0.0)
        ])

    # Formatierungen
    wsB.column_dimensions["A"].width, wsB.column_dimensions["B"].width = 18, 14
    wsB.column_dimensions["C"].width, wsB.column_dimensions["D"].width, wsB.column_dimensions["E"].width = 18, 14, 20
    wsB.column_dimensions["F"].width = 18
    for i in range(2, wsB.max_row + 1):
        wsB.cell(i, 3).number_format = "0.0%"
        wsB.cell(i, 5).number_format = "0.0%"
        bg = _party_color_hex(wsB.cell(i, 1).value)
        wsB.cell(i, 1).fill = PatternFill("solid", fgColor=bg)
        wsB.cell(i, 1).font = _contrast_font(bg)

    # XY-Scatter: X = Diff Posts (F), Y = Second Election (%) (E)
    scatter = ScatterChart()
    scatter.title = "Diff Posts (2-1) vs. Second Election (%)"
    scatter.x_axis.title = "Diff Posts (2-1)"
    scatter.y_axis.title = "Second Election (%)"

    xvalues = Reference(wsB, min_col=6, min_row=2, max_row=wsB.max_row)  # Diff
    yvalues = Reference(wsB, min_col=5, min_row=2, max_row=wsB.max_row)  # Second Election (%)
    series = Series(yvalues, xvalues, title="Parteien")
    scatter.series.append(series)

    wsB.add_chart(scatter, "H2")

    # Datei zurückgeben
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
